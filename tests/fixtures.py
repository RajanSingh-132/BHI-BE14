"""
Shared test fixtures — loads all 8 real datasets once and exposes them
as module-level dicts for use across the test suite.

Ground-truth expected values were computed with pandas directly on the
raw Excel files and are the source of truth for all assertions.
"""

import os
import sys

import openpyxl

# ---------------------------------------------------------------------------
# Ensure project root is on the path regardless of how pytest is invoked
# ---------------------------------------------------------------------------
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

_UPLOAD_DIR = os.path.join(
    os.path.dirname(_PROJECT_ROOT),          # temporaryrajanfe
    os.path.pardir,                           # mnt
    os.path.pardir,                           # busy-relaxed-ride
    "uploads",
)
# Resolve to absolute path
_UPLOAD_DIR = os.path.normpath(
    os.path.join(_PROJECT_ROOT, "..", "..", "uploads")
)


def _load(fname: str) -> list:
    path = os.path.join(_UPLOAD_DIR, fname)
    wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows    = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return [dict(zip(headers, row)) for row in rows]


# ---------------------------------------------------------------------------
# Dataset fixtures (loaded once at module import)
# ---------------------------------------------------------------------------
SF_CAMPAIGN  = _load("Salesforce_Marketing_Campaign_Data.xlsx")
SF_SALES     = _load("Salesforce_Sales_Revenue_From_Marketing_and_Leads.xlsx")
GOOGLE_ADS   = _load("Google_Ads_User_Wise_Sample_Data.xlsx")
MARKETING_REV = _load("Marketing_Revenue_Based_On_Lead_Data.xlsx")
SUGAR_CRM    = _load("sugarcrm_leads_revenueDl.xlsx")
ZOHO_CRM     = _load("Zoho_CRM_Lead_Generation_Sample_Data.xlsx")
SF_LEAD      = _load("Salesforce_Lead_Data_From_Google_Ads.xlsx")
ZOHO_SALES   = _load("Sales_Revenue_From_Zoho_Leads.xlsx")

# ---------------------------------------------------------------------------
# Ground-truth expected values  (computed with pandas on the raw files)
# NEVER derive these from the calculation engine — that would make tests
# circular.  These are the authoritative numbers to assert against.
# ---------------------------------------------------------------------------
EXPECTED = {
    "sf_campaign": {
        "total_leads":       24112,
        "qualified_leads":   15176,
        "converted_leads":    8150,
        "total_cost":       8958211,
        "total_revenue":   52650707,
        "total_profit":    43692496,
        "avg_roi":             605.13,    # AVG(ROI %) per row
        "aggregate_roi":       487.74,    # (SUM Revenue - SUM Cost) / SUM Cost × 100
        "conversion_rate":      33.8,     # SUM(Converted) / SUM(Total) × 100
        "revenue_by_channel": {
            "Email":        14179618,
            "Referral":     11430481,
            "Facebook Ads": 11094831,
        },
        "leads_by_channel_top1": ("Email", 5990),
        "n_rows": 100,
    },
    "sf_sales": {
        "total_leads":        35240,
        "opportunities":      19997,
        "deals_won":          11676,
        "total_cost":       14775690,
        "total_revenue":    97411108,
        "total_profit":     82635418,
        "avg_roi":             709.0,
        "leads_by_channel_top1": ("Google Ads", 8509),
        "n_rows": 120,
    },
    "google_ads": {
        "total_impressions":  990190,
        "total_clicks":        18287,
        "total_cost":         186186.2,
        "total_conversions":    1488,
        "total_revenue":     2706718,
        "avg_cpc":              10.12,
        "aggregate_ctr":         1.8468,
        "aggregate_roi":      1353.77,
        "n_rows": 50,
    },
    "marketing_rev": {
        "total_leads":          8355,
        "qualified_leads":      4722,
        "converted_leads":      2725,
        "total_cost":        3417837,
        "total_revenue":    13938938,
        "total_profit":     10521101,
        "avg_roi":             411.71,
        "n_rows": 50,
    },
    "sugar_crm": {
        "total_revenue":   1550711.06,
        "n_rows": 50,
    },
    "zoho_crm": {
        "total_cost":       73294.7,
        "expected_revenue": 288122.79,
        "n_rows": 100,
    },
    "sf_lead": {
        "avg_cpl":              903.52,
        "expected_revenue":   449072.11,
        "converted_lead_cpl":   867.36,   # CPL filtered to Lead Status = Converted
        "n_rows": 100,
    },
    "zoho_sales": {
        "total_deal_amount": 7391705,
        "total_revenue":     1260445,
        "n_rows": 100,
    },
}
